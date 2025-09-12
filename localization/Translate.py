import requests
import openpyxl

# 🔑 DeepL API Key (buraya kendi keyini gir)
DEEPL_API_KEY = "af27f69a-5c30-4f83-a08e-86807b2175b4:fx"
DEEPL_URL = "https://api-free.deepl.com/v2/translate"

# Excel dosyasını yükle
wb = openpyxl.load_workbook("Localization.xlsx")
ws = wb.active

# Kaynak dil sütunu
source_col = "En"
source_lang = "EN"

# Başlıkları al (1. satırdaki sütun isimleri)
headers = [cell.value for cell in ws[1]]

# Kaynak dil sütununun indexini bul
source_index = headers.index(source_col) + 1  # 1-based index

# Çevrilecek sütunları belirle
language_columns = [col for col in headers if col not in ["Key", source_col]]

# Çeviri fonksiyonu
def translate_text(text, target_lang):
    if text is None or str(text).strip() == "":
        return ""
    response = requests.post(
        DEEPL_URL,
        data={
            "auth_key": DEEPL_API_KEY,
            "text": text,
            "source_lang": source_lang,
            "target_lang": target_lang
        }
    )
    if response.status_code == 200:
        return response.json()["translations"][0]["text"]
    else:
        print(f"Hata ({target_lang}):", response.text)
        return ""

# Satırları dolaş (2. satırdan itibaren)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    en_text = row[source_index - 1].value  # En sütunundaki değer
    if not en_text:
        continue

    for col_name in language_columns:
        col_index = headers.index(col_name)  # sütun indexi
        cell = row[col_index]

        if cell.value is None or str(cell.value).strip() == "":
            target_lang = col_name.upper()  # sütun başlığı → DeepL kodu
            translated = translate_text(str(en_text), target_lang)
            cell.value = translated

# Aynı dosyaya kaydet (stiller korunur)
wb.save("Localization.xlsx")

print("✅ Çeviri tamamlandı. Localization.xlsx güncellendi (stiller korundu).")
